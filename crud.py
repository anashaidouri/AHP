import openpyxl
import pandas as pd
import os

#Loading the excel sheet
def load_workbook(wb_path):
    if os.path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    else:
        return "File not found ..."
    
wb_path = "decision.xlsx"
wb = load_workbook(wb_path)
sheet = wb["Sheet1"]
sheet_obj = wb.active
max_column = sheet_obj.max_column
max_row = sheet_obj.max_row

#Create decision
def Create_decision():
    new_decision = input("\n Enter the decision Name, criteria , subcriteria and alternatives").split(' ')
    sheet.append(new_decision)
    wb.save(wb_path)
    print("decision details added successfully")
    add_more = input("\n Wants to add more decision? Yes/No")
    if add_more.lower() == "yes":
        Create_decision()
        
#View all decisions
def View_all_decision():
    decision_list = pd.read_excel(wb_path)
    print(decision_list)

#Search decision
def search(name):
    for i in range(1,max_row+1):
        if sheet.cell(row=i,column = 1).value == name:
            print("decision found")
            return i

#Display decision
def Display_decision(row):
    for i in range(1,max_column+1):
        cell_obj = sheet_obj.cell(row = row , column = i)
        print(cell_obj.value)

#update decision
def Update_decision(row):
    x = input("\n Enter the decision name, dob, Mail ID").split(' ')
    for col_index,value in enumerate(x,start=1):
        sheet.cell(row =row, column=col_index, value=value)
    print("decision Details Updated successfully")   

#delete decision
def Delete_decision(row):
    sheet.delete_rows(row)
    wb.save(wb_path)
    print("decision deleted successfully")
        

while True:
    print("\n decision Management System")
    print("\n 1.Create decision")
    print("\n 2.View decision")
    print("\n 3.Update decision")
    print("\n 4.Delete decision")
    ch = input("\n Enter the option")
    if ch == '1':
        Create_decision()
    if ch == '2':
        View_all_decision()
    if ch == '3':
        x = input("\n Enter the decision name :")
        row = search(x)
        Display_decision(row)
        y = input("\n wants to edit the decision ? yes/no")
        if y == 'yes':
            Update_decision(row)
    if ch == '4':
        x = input("\n Enter the decision name :")
        row = search(x)
        Display_decision(row)
        y = input("\n wants to delete the decision ? yes/no")
        if y == 'yes':
            Delete_decision(row)
    else:
        break